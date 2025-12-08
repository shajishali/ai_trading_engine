"""
Admin Export Functionality
Provides CSV, Excel, PDF, and JSON export capabilities
"""

import csv
import json
from io import BytesIO, StringIO
from django.http import HttpResponse
from django.utils import timezone
from datetime import datetime
from typing import List, Dict, Any, Optional
import logging

logger = logging.getLogger(__name__)


class BaseExporter:
    """Base class for export functionality"""
    
    def __init__(self, queryset, fields=None, headers=None):
        self.queryset = queryset
        self.fields = fields or []
        self.headers = headers or []
    
    def export(self):
        """Export data - to be implemented by subclasses"""
        raise NotImplementedError


class CSVExporter(BaseExporter):
    """Export data to CSV format"""
    
    def export(self, filename=None):
        """Export queryset to CSV"""
        if filename is None:
            timestamp = timezone.now().strftime('%Y%m%d_%H%M%S')
            filename = f"export_{timestamp}.csv"
        
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        writer = csv.writer(response)
        
        # Write headers
        if self.headers:
            writer.writerow(self.headers)
        elif self.fields:
            writer.writerow(self.fields)
        
        # Write data
        for obj in self.queryset:
            row = []
            for field in self.fields:
                value = self._get_field_value(obj, field)
                row.append(value)
            writer.writerow(row)
        
        return response
    
    def _get_field_value(self, obj, field):
        """Get field value from object"""
        if '__' in field:
            # Handle related fields
            parts = field.split('__')
            value = obj
            for part in parts:
                if value is None:
                    return ''
                value = getattr(value, part, None)
                if callable(value) and not hasattr(value, '__iter__'):
                    value = value()
            return str(value) if value is not None else ''
        else:
            value = getattr(obj, field, '')
            if callable(value) and not hasattr(value, '__iter__'):
                value = value()
            return str(value) if value is not None else ''


class ExcelExporter(BaseExporter):
    """Export data to Excel format"""
    
    def export(self, filename=None):
        """Export queryset to Excel"""
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
        except ImportError:
            logger.warning("openpyxl not installed, falling back to CSV")
            csv_exporter = CSVExporter(self.queryset, self.fields, self.headers)
            return csv_exporter.export(filename.replace('.xlsx', '.csv') if filename else None)
        
        if filename is None:
            timestamp = timezone.now().strftime('%Y%m%d_%H%M%S')
            filename = f"export_{timestamp}.xlsx"
        
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = "Export"
        
        # Style for header
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # Write headers
        headers = self.headers if self.headers else self.fields
        for col_num, header in enumerate(headers, 1):
            cell = worksheet.cell(row=1, column=col_num)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Write data
        for row_num, obj in enumerate(self.queryset, 2):
            for col_num, field in enumerate(self.fields, 1):
                value = self._get_field_value(obj, field)
                cell = worksheet.cell(row=row_num, column=col_num)
                cell.value = value
        
        # Auto-adjust column widths
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width
        
        # Save to response
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        workbook.save(response)
        return response
    
    def _get_field_value(self, obj, field):
        """Get field value from object"""
        if '__' in field:
            parts = field.split('__')
            value = obj
            for part in parts:
                if value is None:
                    return ''
                value = getattr(value, part, None)
                if callable(value) and not hasattr(value, '__iter__'):
                    value = value()
            return value if value is not None else ''
        else:
            value = getattr(obj, field, '')
            if callable(value) and not hasattr(value, '__iter__'):
                value = value()
            return value if value is not None else ''


class JSONExporter(BaseExporter):
    """Export data to JSON format"""
    
    def export(self, filename=None):
        """Export queryset to JSON"""
        if filename is None:
            timestamp = timezone.now().strftime('%Y%m%d_%H%M%S')
            filename = f"export_{timestamp}.json"
        
        data = []
        for obj in self.queryset:
            item = {}
            for field in self.fields:
                value = self._get_field_value(obj, field)
                item[field] = value
            data.append(item)
        
        response = HttpResponse(content_type='application/json')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        response.write(json.dumps(data, indent=2, default=str))
        
        return response
    
    def _get_field_value(self, obj, field):
        """Get field value from object"""
        if '__' in field:
            parts = field.split('__')
            value = obj
            for part in parts:
                if value is None:
                    return None
                value = getattr(value, part, None)
                if callable(value) and not hasattr(value, '__iter__'):
                    value = value()
            return value
        else:
            value = getattr(obj, field, None)
            if callable(value) and not hasattr(value, '__iter__'):
                value = value()
            return value


class PDFExporter(BaseExporter):
    """Export data to PDF format"""
    
    def export(self, filename=None, title="Export Report"):
        """Export queryset to PDF"""
        try:
            from reportlab.lib import colors
            from reportlab.lib.pagesizes import letter, A4
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
            from reportlab.lib.styles import getSampleStyleSheet
            from reportlab.lib.units import inch
        except ImportError:
            logger.warning("reportlab not installed, falling back to CSV")
            csv_exporter = CSVExporter(self.queryset, self.fields, self.headers)
            return csv_exporter.export(filename.replace('.pdf', '.csv') if filename else None)
        
        if filename is None:
            timestamp = timezone.now().strftime('%Y%m%d_%H%M%S')
            filename = f"export_{timestamp}.pdf"
        
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4)
        elements = []
        
        # Title
        styles = getSampleStyleSheet()
        title_para = Paragraph(title, styles['Title'])
        elements.append(title_para)
        elements.append(Spacer(1, 0.2*inch))
        
        # Prepare data
        data = []
        headers = self.headers if self.headers else self.fields
        data.append(headers)
        
        for obj in self.queryset:
            row = []
            for field in self.fields:
                value = self._get_field_value(obj, field)
                row.append(str(value) if value is not None else '')
            data.append(row)
        
        # Create table
        table = Table(data)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
        ]))
        
        elements.append(table)
        
        # Build PDF
        doc.build(elements)
        pdf = buffer.getvalue()
        buffer.close()
        response.write(pdf)
        
        return response
    
    def _get_field_value(self, obj, field):
        """Get field value from object"""
        if '__' in field:
            parts = field.split('__')
            value = obj
            for part in parts:
                if value is None:
                    return ''
                value = getattr(value, part, None)
                if callable(value) and not hasattr(value, '__iter__'):
                    value = value()
            return str(value) if value is not None else ''
        else:
            value = getattr(obj, field, '')
            if callable(value) and not hasattr(value, '__iter__'):
                value = value()
            return str(value) if value is not None else ''


def export_queryset(queryset, format='csv', fields=None, headers=None, filename=None):
    """
    Export a queryset to various formats
    
    Args:
        queryset: Django queryset to export
        format: Export format ('csv', 'excel', 'json', 'pdf')
        fields: List of field names to export
        headers: List of header names (optional)
        filename: Output filename (optional)
    
    Returns:
        HttpResponse with exported data
    """
    exporters = {
        'csv': CSVExporter,
        'excel': ExcelExporter,
        'xlsx': ExcelExporter,
        'json': JSONExporter,
        'pdf': PDFExporter,
    }
    
    exporter_class = exporters.get(format.lower(), CSVExporter)
    exporter = exporter_class(queryset, fields, headers)
    
    if filename and not filename.endswith(f'.{format}'):
        filename = f"{filename}.{format}"
    
    return exporter.export(filename)













